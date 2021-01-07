from openpyxl import load_workbook
from openpyxl.styles import Font, colors


class Write_excel(object):
    '''修改Excel数据'''

    def __init__(self,filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  #激活sheet

    def write(self,row_n,col_n,value):
        '''写入数据，如（2，3，hello），第二行第三列写入数据hello'''
        ft = Font(color=colors.RED,size=12,bold=True)
        #判断值为错误时添加字体样式
        if value in ['fail','error'] or col_n == 12:
            self.ws.cell(row_n,col_n).font = ft
        if value == 'pass':
            ft = Font(color=colors.GREEN)
            self.ws.cell(row_n,col_n).font = ft
        self.ws.cell(row_n,col_n).value = value
        self.wb.save(self.filename)
